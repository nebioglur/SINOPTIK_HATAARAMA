import pandas as pd
import os
import re
from collections import Counter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import kurallar
def raporu_excel_olarak_kaydet(birlesik, yil, ay, okuma_raporu, cikti_klasoru=None, islenen_dosyalar=None):
    
    # Detaylı raporda boş hücreleri temiz bırak (çizgi koyma)
    def bos_hucreleri_temizle(df):
        df_out = df.copy()
        df_out = df_out.fillna("")
        for col in df_out.columns:
            if df_out[col].dtype == 'object':
                mask = (df_out[col].astype(str).str.strip() == "-") | (df_out[col].astype(str).str.strip() == "nan")
                df_out.loc[mask, col] = ""
        return df_out

    if not cikti_klasoru:
        cikti_klasoru = os.path.join(os.environ["USERPROFILE"], "Desktop")
    
    # Dosya ismini belirle (Eğer dosya açıksa üzerine yazamaz, yeni isim türet)
    base_filename = f"DENETIM_{yil}_{ay:02d}"
    cikti = os.path.join(cikti_klasoru, f"{base_filename}.xlsx")
    
    counter = 1
    while True:
        try:
            # Dosya kilitli mi kontrol et (PermissionError [Errno 13] yakalamak için)
            with open(cikti, 'a'): pass
            break
        except PermissionError:
            cikti = os.path.join(cikti_klasoru, f"{base_filename} ({counter}).xlsx")
            counter += 1
            
    # Görünümü iyileştirmek için tamamen boş olan gereksiz sütunları kaldır
    birlesik = birlesik.dropna(axis=1, how='all')

    # Tekrar eden veya pandas'ın otomatik isimlendirdiği kopya sütunları (Örn: Personel_1) temizle
    silinecek_sutunlar = []
    temel_isimler = set()
    for col in birlesik.columns:
        # "_1", "_2" gibi otomatik kopya eklerini temizle
        temel_ad = re.sub(r'_\d+$', '', str(col)).strip().upper()
        if temel_ad in temel_isimler:
            silinecek_sutunlar.append(col)
        else:
            temel_isimler.add(temel_ad)
            
    if silinecek_sutunlar:
        birlesik = birlesik.drop(columns=silinecek_sutunlar)

    hatalar_df = birlesik[birlesik["DURUM"] != "Hata Yok"]

    # --- İSTATİSTİKLER ---
    # 1. Hata Kodları
    tum_kodlar = []
    if not hatalar_df.empty:
        for k in hatalar_df["HATA KODU"].dropna():
            tum_kodlar.extend([x.strip() for x in str(k).split(",") if x.strip()])
    kod_sayilari = Counter(tum_kodlar)
    
    # 1. ÖZET Sayfası ve Grafik İçin Sadece Gerçekleşen Hatalar
    df_kod_ist = pd.DataFrame(kod_sayilari.most_common(), columns=["Hata Kodu", "Adet"])
    
    tum_kurallar_listesi = []
        
    # 1. Sözlükte tanımlı olan tüm güncel kuralları ekle
    for k_kod, k_aciklama in kurallar.HATA_SOZLUGU.items():
        adet = kod_sayilari.get(k_kod, 0)
        durum = "BAŞARISIZ ❌" if adet > 0 else "BAŞARILI ✅"
        tum_kurallar_listesi.append({"Kural Kodu": k_kod, "Durum": durum, "Hata Sayısı": adet, "Açıklama": k_aciklama})
        
    # 2. Sözlükte olmayan ancak analizde tespit edilen sistem içi çapraz kontrolleri ekle
    mevcut_kodlar = [x["Kural Kodu"] for x in tum_kurallar_listesi]
    for k_kod, adet in kod_sayilari.items():
        if k_kod not in mevcut_kodlar:
            durum = "BAŞARISIZ ❌" if adet > 0 else "BAŞARILI ✅"
            k_aciklama = kurallar.HATA_SOZLUGU.get(k_kod, "Sistem İçi Dinamik Çapraz Kontrol")
            if k_kod == "Veri Yok":
                k_aciklama = "İlgili ana/ara sinoptik saatinde rasat verisi bulunamadı (Tüm zorunlu parametreler eksik)."
            elif k_kod == "Ara Rasat":
                k_aciklama = "Sadece METAR bulunur, SİNOPTİK beklenmez."
            tum_kurallar_listesi.append({"Kural Kodu": k_kod, "Durum": durum, "Hata Sayısı": adet, "Açıklama": k_aciklama})
            
    df_tum_kurallar = pd.DataFrame(tum_kurallar_listesi)
    df_tum_kurallar["Sira"] = df_tum_kurallar["Kural Kodu"].apply(lambda x: int(re.search(r'\d+', str(x)).group()) if re.search(r'\d+', str(x)) else 9999)
    df_tum_kurallar = df_tum_kurallar.sort_values(by=["Sira", "Kural Kodu"]).drop(columns=["Sira"])

    # 2. Günlük Hata
    if "Tarih" in hatalar_df.columns:
        df_gun_ist = hatalar_df["Tarih"].value_counts().reset_index()
        df_gun_ist.columns = ["Tarih", "Hata Sayısı"]
        df_gun_ist = df_gun_ist.sort_values(by="Tarih")
    else:
        df_gun_ist = pd.DataFrame()

    # 3. Personel Hata İstatistikleri
    df_personel_ist = pd.DataFrame()
    personel_cols = [c for c in hatalar_df.columns if "Personel" in c]
    if personel_cols:
        tum_personel = []
        for col in personel_cols:
            tum_personel.extend(hatalar_df[col].dropna().astype(str).tolist())
        
        if tum_personel:
            p_counts = Counter(tum_personel)
            df_personel_ist = pd.DataFrame(p_counts.most_common(), columns=["Personel", "Hata Sayısı"])

    # 4. Genel
    genel_bilgiler = [
        {"Bilgi": "Rapor Dönemi", "Değer": f"{ay}/{yil}"},
        {"Bilgi": "Toplam Kayıt", "Değer": len(birlesik)},
        {"Bilgi": "Hatalı Kayıt", "Değer": len(hatalar_df)},
        {"Bilgi": "Hata Oranı", "Değer": f"%{(len(hatalar_df)/len(birlesik)*100):.1f}" if len(birlesik) > 0 else "0"},
    ]
    
    if islenen_dosyalar:
        genel_bilgiler.append({"Bilgi": "İşlenen SİNOPTİK", "Değer": islenen_dosyalar.get("sinoptik", "-")})
        genel_bilgiler.append({"Bilgi": "İşlenen METAR", "Değer": islenen_dosyalar.get("metar", "-")})
        
    df_genel = pd.DataFrame(genel_bilgiler)

    with pd.ExcelWriter(cikti, engine="openpyxl") as writer:
        # ÖZET Sayfası
        df_genel.to_excel(writer, sheet_name="OZET", index=False, startrow=5, startcol=1)
        df_kod_ist.to_excel(writer, sheet_name="OZET", index=False, startrow=5, startcol=4)
        df_gun_ist.to_excel(writer, sheet_name="OZET", index=False, startrow=5, startcol=7)
        if not df_personel_ist.empty:
            df_personel_ist.to_excel(writer, sheet_name="OZET", index=False, startrow=5, startcol=10)

        ws_ozet = writer.sheets["OZET"]
        
        # 1. KURUMSAL BANNER (Başlık)
        ws_ozet.merge_cells("B2:L4")
        banner_cell = ws_ozet["B2"]
        banner_cell.value = f"MGM KARDELEN - {ay:02d}/{yil} DENETİM RAPORU GÖSTERGE PANELİ"
        banner_cell.font = Font(name="Segoe UI", size=18, bold=True, color="FFFFFF")
        banner_cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        banner_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Arka plan ızgaralarını kapat
        ws_ozet.sheet_view.showGridLines = False
        
        ws_ozet["B5"] = "GENEL DURUM"
        ws_ozet["E5"] = "HATA TÜRÜ DAĞILIMI"
        ws_ozet["H5"] = "GÜNLÜK HATA DAĞILIMI"
        ws_ozet["K5"] = "PERSONEL HATA DAĞILIMI"
        for cell in ["B5", "E5", "H5", "K5"]:
            ws_ozet[cell].font = Font(name="Segoe UI", bold=True, size=12, color="2C3E50")
            ws_ozet[cell].alignment = Alignment(horizontal="center", vertical="center")

        # 📊 GRAFİK EKLEME
        if not df_kod_ist.empty:
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "En Sık Yapılan 5 Hata"
            chart.y_axis.title = "Adet"
            chart.x_axis.title = "Hata Kodu"
            chart.legend = None

            limit_kod = min(5, len(df_kod_ist))
            data = Reference(ws_ozet, min_col=6, min_row=6, max_row=6+limit_kod, max_col=6)
            cats = Reference(ws_ozet, min_col=5, min_row=7, max_row=6+limit_kod)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws_ozet.add_chart(chart, "N6")

        # 📊 GRAFİK EKLEME (Personel Hataları)
        if not df_personel_ist.empty:
            chart_pers = BarChart()
            chart_pers.type = "col"
            chart_pers.style = 12
            chart_pers.title = "En Çok Hata Yapan 5 Personel"
            chart_pers.y_axis.title = "Hata Sayısı"
            chart_pers.x_axis.title = "Personel"
            chart_pers.legend = None

            limit_pers = min(5, len(df_personel_ist))
            data_pers = Reference(ws_ozet, min_col=12, min_row=6, max_row=6+limit_pers, max_col=12)
            cats_pers = Reference(ws_ozet, min_col=11, min_row=7, max_row=6+limit_pers)
            chart_pers.add_data(data_pers, titles_from_data=True)
            chart_pers.set_categories(cats_pers)
            ws_ozet.add_chart(chart_pers, "W6")
            
        # ÖZET Sayfası Veri Başlıklarını Ortala ve Boya
        for cell in ws_ozet[6]:
            if cell.value:
                cell.font = Font(name="Segoe UI", bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="16A085", end_color="16A085", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
        # Sütun Genişliklerini Ayarla
        for col_letter in ['B', 'C', 'E', 'F', 'H', 'I', 'K', 'L']:
            ws_ozet.column_dimensions[col_letter].width = 15

        # 📊 YENİ: TÜM KURALLAR (h1-h267) Sayfası
        df_tum_kurallar.to_excel(writer, sheet_name="TÜM_KURALLAR", index=False)
        ws_tum_kurallar = writer.sheets["TÜM_KURALLAR"]
        ws_tum_kurallar.column_dimensions["A"].width = 15
        ws_tum_kurallar.column_dimensions["B"].width = 20
        ws_tum_kurallar.column_dimensions["C"].width = 15
        ws_tum_kurallar.column_dimensions["D"].width = 100
        for cell in ws_tum_kurallar[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0288D1", end_color="0288D1", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Sütunları mantıksal gruplara ayırarak sırala
        oncelikli_sutunlar = [
            "sayfa", "gmt", "DURUM", "HATA KODU", "AÇIKLAMA", "HATALI KOD", "TAVSİYE KOD",
            "SİNOPTİK - Şifreli Mesaj", "METAR - Şifreli Mesaj", "bulten"
        ]
        
        # Kalan sütunları (detay parametreleri) al
        kalan_sutunlar = [col for col in birlesik.columns if col not in oncelikli_sutunlar]
        yeni_sira = [col for col in oncelikli_sutunlar if col in birlesik.columns] + kalan_sutunlar
        birlesik = birlesik[yeni_sira]
        
        # Detay parametrelerinin başladığı sütun indeksini bul (Gruplama için)
        detay_baslangic_idx = len([col for col in oncelikli_sutunlar if col in birlesik.columns]) + 1
        
        birlesik_export = bos_hucreleri_temizle(birlesik)
        
        # AÇIKLAMA sütununu artık ÇIKARMIYORUZ, kullanıcı bunu görmek istiyor.
        birlesik_export.to_excel(writer, sheet_name="Detaylı_Rapor", index=False, startrow=1)
        
        if not hatalar_df.empty:
            hatalar_export = bos_hucreleri_temizle(hatalar_df)
            hatalar_export.to_excel(writer, sheet_name="SADECE_HATALAR", index=False, startrow=1)

            # 🔹 HATA KODUNA GÖRE AYRI SAYFALAR
            tum_hatalar = set()
            for kodlar in hatalar_df["HATA KODU"]:
                if pd.isna(kodlar): continue
                for kod in str(kodlar).split(","):
                    tum_hatalar.add(kod.strip())

            for hata in sorted(tum_hatalar):
                def hata_var_mi(val):
                    if pd.isna(val): return False
                    kodlar = [k.strip() for k in str(val).split(",")]
                    return hata in kodlar

                df_ozel = hatalar_df[hatalar_df["HATA KODU"].apply(hata_var_mi)]
                
                if not df_ozel.empty:
                    sheet_name = re.sub(r'[\\/*?:\[\]]', '', str(hata))[:31]
                    try:
                        df_ozel_export = bos_hucreleri_temizle(df_ozel)
                        df_ozel_export.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
                    except: pass

        # STİLLER (with bloğu içine alındı)
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        
        # --- ÖZELLEŞTİRİLEBİLİR RENK PALETİ (MODERN FLAT DESIGN) ---
        sinoptik_header_fill = PatternFill(start_color="16A085", end_color="16A085", fill_type="solid") # Modern Zarif Turkuaz/Yeşil
        metar_header_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid") # Modern Belize Hole Mavi
        neutral_header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid") # Koyu Lacivert/Gri (Dark Slate)
        
        # Hata Türü Renkleri
        wmo_error_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid") # Soft Kırmızı/Pembe
        uyum_error_fill = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid") # Soft Turuncu
        missing_fill = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid") # Soft Sarı
        ok_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid") # Soft Yeşil (Hata Yok)
        
        # Hata Türü Yazı Tipleri (Kalın ve okunaklı)
        wmo_error_font = Font(name="Segoe UI", size=10, bold=True, color="C0392B") # Koyu Kırmızı yazı
        uyum_error_font = Font(name="Segoe UI", size=10, bold=True, color="D35400") # Koyu Turuncu yazı
        missing_font = Font(name="Segoe UI", size=10, color="7F8C8D", italic=True) # Gri italik yazı
        ok_font = Font(name="Segoe UI", size=10, bold=True, color="27AE60") # Koyu Yeşil yazı
        normal_font = Font(name="Segoe UI", size=10, color="2C3E50") # Ana metin (Siyah yerine Koyu Lacivert/Gri)
        
        zebra_fill = PatternFill(start_color="F8F9F9", end_color="F8F9F9", fill_type="solid") # Çok açık soft gri (Zebra)
        
        # Çerçeve Stilleri (Daha ince ve açık renk)
        thin_border = Border(left=Side(style='thin', color='ECF0F1'),
                             right=Side(style='thin', color='ECF0F1'),
                             top=Side(style='thin', color='ECF0F1'),
                             bottom=Side(style='thin', color='ECF0F1'))
                             
        error_border = Border(left=Side(style='medium', color='E74C3C'),
                              right=Side(style='medium', color='E74C3C'),
                              top=Side(style='medium', color='E74C3C'),
                              bottom=Side(style='medium', color='E74C3C'))
                             
        # Sütun Arkaplanları (Artık çok çok hafif renkler)
        sinoptik_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") # Temiz Beyaz
        sinoptik_zebra = zebra_fill
        
        metar_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") # Temiz Beyaz
        metar_zebra = zebra_fill
                             
        hata_sutun_map = {
            "h48": "Sıcaklık (T)", "h53": "İşba (Td)", "h187": "Maks. Sıcaklık (Tx)", "h193": "Min. Sıcaklık (Tn)",
            "h55": "İstasyon Basıncı (P)", "h58": "Deniz Basıncı (P0)", "h64": "Basınç Değişimi (ppp)",
            "h44": "Rüzgar Hızı (ff)", "h37": "Rüzgar Yönü (dd)", "h38": "Rüzgar", "h39": "Rüzgar", "h43": "Rüzgar",
            "h70": "Yağış Miktarı (RRR)", "h71": "Yağış Miktarı (RRR)", "h74": "Yağış Süresi (tR)", "h4": "Yağış Miktarı (RRR)",
            "h22": "Görüş (VV)", "h17": "Bulut Yük. (h)",
            "h26": "Toplam Bulut (N)", "h27": "Toplam Bulut (N)", "h34": "Toplam Bulut (N)", "h120": "Toplam Bulut (N)",
            "h3": "İndikatör (ir)", "h5": "İndikatör (ir)", "h8": "İndikatör", "h11": "İndikatör (ix)", "h12": "İndikatör (ix)", "h13": "İndikatör (ix)",
            "h252": "Sıcaklık (T)", "h253": "İşba (Td)", "h254": "İstasyon Basıncı (P)", "h255": "Deniz Basıncı (P0)", "h257": "Toplam Bulut (N)",
            "h270": "Sıcaklık (T)", "h271": "İşba (Td)", "h272": "Bağıl Nem (%)", "h273": "Rüzgar Hızı (ff)", "h274": "Rüzgar Yönü (dd)",
            "h275": "Rüzgar", "h276": "910 Grubu (Hamle)", "h277": "911 Grubu (Hamle)", "h278": "Toplam Bulut (N)",
            "h279": "Bulut Yük. (h)", "h280": "Alçak Bulut (CL)", "h281": "Görüş (VV)", "h282": "Görüş (VV)",
            "h283": "Görüş (VV)", "h284": "Bulut Yük. (h)", "h285": "Alçak Bulut (CL)",
            "h290": "924 Grubu", "h304": "Sıcaklık (T)", "h305": "Sıcaklık (T)", "h306": "Deniz Basıncı (P0)",
            "h309": "Geçmiş Hava 1 (W1)", "h311": "Bulut Yük. (h)", "h312": "Görüş (VV)",
            "h313": "Halihazır Hava (ww)", "h314": "Alçak/Orta Bulut (Nh)", "h315": "Yağış Miktarı (RRR)",
            "h316": "Yağış Miktarı (RRR)", "h317": "Yağış Miktarı (RRR)", "h318": "Alçak Bulut (CL)",
            "h319": "Halihazır Hava (ww)", "h320": "Halihazır Hava (ww)", "h321": "Halihazır Hava (ww)",
            "h322": "Geçmiş Hava 1 (W1)", "h323": "Görüş (VV)", "h324": "Bulut Yük. (h)", "h325": "Alçak Bulut (CL)",
            "h334": "Halihazır Hava (ww)", "h335": "Yağış Miktarı (RRR)", "h340": "910 Grubu (Hamle)",
            "h341": "960 Grubu (Hadise)", "h342": "960 Grubu (Hadise)", "h343": "932 Grubu (Taze Kar)",
            "h344": "932 Grubu (Taze Kar)", "h345": "932 Grubu (Taze Kar)", "h346": "931 Grubu (Kar)",
            "h347": "931 Grubu (Kar)", "h353": "Alçak Bulut (CL)", "h354": "Orta Bulut (CM)",
            "h355": "Yüksek Bulut (CH)", "h356": "Alçak Bulut (CL)", "h357": "Alçak/Orta Bulut (Nh)",
            "h358": "Toplam Bulut (N)", "h359": "924 Grubu", "h360": "924 Grubu", "h361": "Geçmiş Hava 1 (W1)",
            "h362": "SİNOPTİK - Şifreli Mesaj", "h364": "Halihazır Hava (ww)", "h365": "Halihazır Hava (ww)",
            "h366": "Halihazır Hava (ww)", "h367": "Halihazır Hava (ww)", "h368": "Halihazır Hava (ww)",
            "h369": "Halihazır Hava (ww)", "h370": "Halihazır Hava (ww)", "h371": "Halihazır Hava (ww)",
            "h372": "Halihazır Hava (ww)", "h373": "Halihazır Hava (ww)", "h374": "932 Grubu (Taze Kar)",
            "h375": "Rüzgar Yönü (dd)", "h376": "Yağış Süresi (tR)", "h377": "Görüş (VV)"
        }

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            if sheet_name == "OZET": continue

            ws["A1"] = "RENK LEJANTI:"
            ws["A1"].font = Font(name="Segoe UI", bold=True)
            ws["B1"] = "WMO Hatası (Kırmızı)"
            ws["B1"].fill = wmo_error_fill
            ws["B1"].font = normal_font
            ws["B1"].alignment = Alignment(horizontal='center')
            ws["C1"] = "Uyumsuzluk/Uyarı (Turuncu)"
            ws["C1"].fill = uyum_error_fill
            ws["C1"].font = normal_font
            ws["C1"].alignment = Alignment(horizontal='center')
            ws["D1"] = "Veri Yok/Eksik (Sarı)"
            ws["D1"].fill = missing_fill
            ws["D1"].font = normal_font
            ws["D1"].alignment = Alignment(horizontal='center')
            ws["E1"] = "Hata Yok (Yeşil)"
            ws["E1"].fill = ok_fill
            ws["E1"].font = normal_font
            ws["E1"].alignment = Alignment(horizontal='center')
            
            # Daha modern ve temiz bir görünüm için Excel'in varsayılan ızgara çizgilerini kapat
            ws.sheet_view.showGridLines = False

            durum_col_idx = None
            hata_kodu_col_idx = None

            # Başlıklar 2. satırda (startrow=1 olduğu için)
            for cell in ws[2]:
                cell.font = header_font
                header_text = str(cell.value).upper() if cell.value else ""
                if "SİNOPTİK" in header_text:
                    cell.fill = sinoptik_header_fill
                elif "METAR" in header_text:
                    cell.fill = metar_header_fill
                else:
                    cell.fill = neutral_header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                if cell.value and str(cell.value) == "DURUM": durum_col_idx = cell.column
                if cell.value and str(cell.value) == "HATA KODU": hata_kodu_col_idx = cell.column

            if durum_col_idx and hata_kodu_col_idx:
                ws.freeze_panes = "C3" # İlk iki sütunu (Tarih, Saat) ve başlıkları dondur
                ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"
                ws.row_dimensions[2].height = 35 # Başlıkların rahat okunması için
                
                # Sütunları grupla ve daraltılmış (gizlenmiş) olarak ayarla
                if ws.max_column >= detay_baslangic_idx:
                    start_col_letter = get_column_letter(detay_baslangic_idx)
                    end_col_letter = get_column_letter(ws.max_column)
                    ws.column_dimensions.group(start_col_letter, end_col_letter, hidden=True)
                    
                header_cells = [str(c.value).upper() if c.value else "" for c in ws[2]]
                
                # Sütun bazlı arkaplan renklerini belirle
                col_base_fills = []
                for header in header_cells:
                    if "SİNOPTİK" in header:
                        col_base_fills.append((sinoptik_fill, sinoptik_zebra))
                    elif "METAR" in header:
                        col_base_fills.append((metar_fill, metar_zebra))
                    else:
                        col_base_fills.append((None, zebra_fill))
                
                # Veriler 3. satırdan başlıyor
                for row_idx, row in enumerate(ws.iter_rows(min_row=3), start=3):
                    durum_val = row[durum_col_idx-1].value
                    hata_val = str(row[hata_kodu_col_idx-1].value or "")
                    
                    is_zebra = (row_idx % 2 == 0)
                    
                    for col_idx_cell, cell in enumerate(row):
                        cell.border = thin_border
                        
                        base_f, zebra_f = None, zebra_fill
                        if col_idx_cell < len(col_base_fills):
                            base_f, zebra_f = col_base_fills[col_idx_cell]
                            
                        if is_zebra and zebra_f:
                            cell.fill = zebra_f
                        elif not is_zebra and base_f:
                            cell.fill = base_f
                            
                        # Standart modern font uygula
                        cell.font = normal_font

                    if durum_val == "Veri Yok" or "Veri Yok" in hata_val:
                        for cell in row:
                            cell.fill = missing_fill
                            cell.font = missing_font
                    elif durum_val == "Hata Yok":
                        row[durum_col_idx-1].fill = ok_fill
                        row[durum_col_idx-1].font = ok_font
                    elif durum_val not in ["Hata Yok", "Ara Rasat"] and durum_val is not None:
                        # Dinamik Özel Uyarılar için Turuncu, Standart Kurallar (h1, h2) için Kırmızı
                        fill_color = wmo_error_fill
                        font_color = wmo_error_font
                        
                        # "Mantık", "Çapraz", "Uyumsuzluk" vb. içeriyorsa ve içinde standart h1-h267 kuralı yoksa turuncu yap
                        if re.search(r'(UYUM|TUTARSIZLIK|MANTIK|ÇAPRAZ|DİKKAT)', hata_val.upper()):
                            if not re.search(r'\bh\d+\b', hata_val):
                                fill_color = uyum_error_fill
                                font_color = uyum_error_font
                                
                        # Hata kodu hücresini boya ve vurgula
                        row[hata_kodu_col_idx-1].fill = fill_color
                        row[hata_kodu_col_idx-1].font = font_color
                        
                        # Hatalı hücrenin kendisini de boya
                        kodlar = [k.strip().split()[0] for k in hata_val.split(",")] 
                        for kod in kodlar:
                            target = hata_sutun_map.get(kod)
                            if target:
                                for idx, header in enumerate(header_cells):
                                    if target.upper() in header:
                                        row[idx].fill = fill_color
                                        row[idx].font = font_color
                                        row[idx].border = error_border

            for column_cells in ws.columns:
                max_length = 0
                col_letter = get_column_letter(column_cells[0].column)
                
                # Kolon başlığını kontrol et (Başlıklar 2. satırda yer alır)
                header_val = str(column_cells[1].value).upper() if len(column_cells) > 1 and column_cells[1].value else ""
                is_rasatlar = "RASATLAR" in header_val or "BÜLTEN" in header_val or "HAM VERİ" in header_val
                
                is_long_text = "AÇIKLAMA" in header_val or "MESAJ" in header_val
                for cell in column_cells:
                    try:
                        if cell.row > 2: # Sadece veri satırları
                            if is_rasatlar or is_long_text:
                                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                            else:
                                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                        if cell.value:
                            cell_len = len(str(cell.value))
                            if cell_len > max_length: max_length = cell_len
                    except: pass

                if is_rasatlar:
                    ws.column_dimensions[col_letter].width = 40
                elif is_long_text:
                    ws.column_dimensions[col_letter].width = 50
                else:
                    ws.column_dimensions[col_letter].width = min(max(max_length + 3, 12), 30)

    return cikti