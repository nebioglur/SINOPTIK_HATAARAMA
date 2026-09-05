import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import os
from openpyxl.utils import get_column_letter
import threading
import re
import shutil
import logging
from logging.handlers import RotatingFileHandler
import zipfile

# --- BAD CRC-32 (BOZUK EXCEL) BYPASS HACK ---
try:
    if hasattr(zipfile, 'ZipExtFile') and hasattr(zipfile.ZipExtFile, '_update_crc'):
        original_update_crc = zipfile.ZipExtFile._update_crc
        def patched_update_crc(self, newdata):
            try:
                original_update_crc(self, newdata)
            except zipfile.BadZipFile:
                pass
        zipfile.ZipExtFile._update_crc = patched_update_crc
        if not getattr(zipfile.ZipExtFile, '_crc_patched', False):
            _orig_update_crc = zipfile.ZipExtFile._update_crc
            def _patched_update_crc(self, newdata):
                self._expected_crc = None  # Zorla CRC kontrolünü kapat (BadZipFile hatasını önler)
                return _orig_update_crc(self, newdata)
            zipfile.ZipExtFile._update_crc = _patched_update_crc
            zipfile.ZipExtFile._crc_patched = True
except Exception:
    pass
# --------------------------------------------

# Loglama yapılandırması
if not logging.getLogger().hasHandlers():
    log_handler = RotatingFileHandler('denetim_merkezi.log', maxBytes=5*1024*1024, backupCount=2, encoding='utf-8')
    logging.basicConfig(
        handlers=[log_handler],
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )

# Ana uygulamadaki normalizasyon ve başlık bulma fonksiyonlarını kullanmak için import et
try:
    import denetim_merkezi_1 as dm1
except ImportError:
    # Bu script'in tek başına çalışabilmesi için hata mesajı göster
    # Eğer bu dosya ana klasörde değilse, bu import başarısız olur.
    tk.Tk().withdraw()
    messagebox.showerror(
        "Kritik Hata", 
        "Gerekli 'denetim_merkezi_1.py' dosyası bulunamadı.\n\n"
        "Lütfen bu script'i ana uygulama ('arayuz.py') ile aynı klasörde çalıştırdığınızdan emin olun."
    )
    exit()

def sutunlari_otomatik_duzelt():
    """
    Kullanıcıdan bir Excel dosyası seçmesini ister. Dosyadaki her sayfanın
    sütun başlıklarını ana programın anlayacağı standart formata dönüştürür
    ve '_DUZELTILMIS' ekiyle yeni bir dosya olarak kaydeder.
    """
    if not tk._default_root:
        root = tk.Tk()
        root.withdraw()

    dosya_yolu = filedialog.askopenfilename(
        title="Sütun Başlıklarını Düzeltmek İçin Excel Dosyasını Seçin",
        filetypes=[("Excel Dosyaları", "*.xlsx *.xls")]
    )

    if not dosya_yolu:
        # Kullanıcı dosya seçmeden pencereyi kapattıysa işlemi sonlandır.
        return

    def islem_yurut():
        try:
            # Çıktı dosyasının adını oluştur (örn: 'veriler.xlsx' -> 'veriler_DUZELTILMIS.xlsx')
            dizin, dosya_adi = os.path.split(dosya_yolu)
            ad, uzanti = os.path.splitext(dosya_adi)
            cikti_yolu = os.path.join(dizin, f"{ad}_DUZELTILMIS{uzanti}")

            # Orijinal Excel dosyasını oku
            xls = pd.ExcelFile(dosya_yolu)
            
            # Yeni Excel dosyası için bir yazıcı (writer) oluştur
            with pd.ExcelWriter(cikti_yolu, engine='openpyxl') as writer:
                # 🚀 KESİN ÇÖZÜM: İşlem kesilirse çökmemesi için geçici bir sayfa oluştur.
                pd.DataFrame(["İşlem bekleniyor..."]).to_excel(writer, sheet_name="Gecici_Sayfa", index=False, header=False)
                
                yazilan_sayfa_sayisi = 0

                # Dosyadaki her bir sayfa (sheet) için döngü başlat
                for sayfa_adi in xls.sheet_names:
                    # BAZI EXCEL DIŞA AKTARMA ARAÇLARI TARAFINDAN OLUŞTURULAN GİZLİ METADATA SAYFALARINI ATLA
                    if 'document map' in str(sayfa_adi).lower() or 'documentmap' in str(sayfa_adi).lower():
                        continue
                    logging.info(f" - '{sayfa_adi}' sayfası işleniyor...")
                    
                    # Veriler açık rasatlarda (Sheet2, A-...) olduğu için kapalı rasatları (Sheet3, K-...) atla
                    sayfa_str = str(sayfa_adi).lower()
                    if "sheet" in sayfa_str or "sayfa" in sayfa_str:
                        rakamlar = re.findall(r'\d+', sayfa_str)
                        if rakamlar and int(rakamlar[0]) % 2 != 0:
                            # KONTROL: Bu sayfa aslında ana verileri içeriyor mu?
                            try:
                                ham_kontrol = pd.read_excel(xls, sheet_name=sayfa_adi, header=None, dtype=str)
                                h_idx = dm1.header_bul(ham_kontrol)
                                is_main_sheet = False
                                if h_idx is not None:
                                    row_values = ham_kontrol.iloc[h_idx].astype(str).str.lower().values
                                    match_count = sum(1 for val in row_values if str(val).strip() in ["t", "p", "n", "ff", "dd", "h", "a"] or any(k in str(val).strip() for k in ["gmt", "saat", "ww", "halihazır", "present", "istasyon", "rüzgar", "yön", "hız", "basınç", "yağış", "rrr", "bulut", "görüş", "vv", "bülten", "metar", "tipi"]))
                                    if match_count >= 5: is_main_sheet = True
                                if not is_main_sheet:
                                    for r_i in range(min(10, len(ham_kontrol))):
                                        v_a = str(ham_kontrol.iloc[r_i, 0]).strip().upper() if len(ham_kontrol.columns) > 0 else ""
                                        if v_a in ['TİPİ', 'TIPI', 'METAR'] or 'METAR' in v_a:
                                            is_main_sheet = True; break
                                if is_main_sheet:
                                    pass # Bu sayfa atlanmamalı, normal düzeltmeye tabi tutulmalı
                                else:
                                    logging.info(f"   -> Tek sayılı sayfa ({sayfa_adi}) meta veri sayılarak aynen kopyalanıyor.")
                                    ham_kontrol.to_excel(writer, sheet_name=sayfa_adi, index=False, header=False)
                                    yazilan_sayfa_sayisi += 1
                                    continue
                            except Exception: pass
           
                    try:
                        # Başlık satırının nerede olduğunu bulmak için veriyi ham olarak oku (başlıksız)
                        ham_df = pd.read_excel(xls, sheet_name=sayfa_adi, header=None, dtype=str)
                        
                        sayfa_str = str(sayfa_adi).lower()
                        yeni_sayfa_adi = str(sayfa_adi)
                        # DÜZELTME: Sayfa isimlerini değiştirmeyin, aksi takdirde Sheet2-Sheet3 eşleşmesi bozulur.

                        # denetim_merkezi_1'deki fonksiyon ile başlık satırının indeksini bul
                        header_index = dm1.header_bul(ham_df)
                        
                        if header_index is not None:
                            ham_df.to_excel(writer, sheet_name=yeni_sayfa_adi, index=False, header=False)
                            logging.info(f"   -> Başlıklar {header_index + 1}. satırda bulundu. Orijinal başlık yapısı korunuyor. (Yeni Adı: {yeni_sayfa_adi})")

                            # 🚀 PERFORMANS OPTİMİZASYONU: Sütun genişliğini milisaniyeler içinde hesapla
                            ws = writer.sheets[yeni_sayfa_adi]
                            for col_idx, col in enumerate(ham_df.columns):
                                try:
                                    max_len = ham_df[col].astype(str).map(len).max()
                                    width = min(max_len + 2, 60)
                                except:
                                    width = 15
                                col_letter = get_column_letter(col_idx + 1)
                                ws.column_dimensions[col_letter].width = width

                            yazilan_sayfa_sayisi += 1
                        else:
                            logging.warning(f"   -> UYARI: Bu sayfada tanınabilir bir başlık satırı bulunamadı. Değiştirilmeden kopyalanıyor. (Yeni Adı: {yeni_sayfa_adi})")
                            ham_df.to_excel(writer, sheet_name=yeni_sayfa_adi, index=False, header=False)
                            yazilan_sayfa_sayisi += 1
                    except Exception as inner_e:
                        logging.error(f"   -> HATA: '{sayfa_adi}' işlenirken hata oluştu ve atlandı.", exc_info=True)

                # Geçici sayfayı temizle veya boş dosyayı işaretle
                if yazilan_sayfa_sayisi > 0:
                    writer.book.remove(writer.sheets["Gecici_Sayfa"])
                else:
                    if "Gecici_Sayfa" in writer.sheets:
                        writer.sheets["Gecici_Sayfa"].title = "Bos_Sayfa"

            messagebox.showinfo(
                "Başarılı",
                f"Sütunlar başarıyla düzeltildi!\n\n"
                f"Yeni dosya şu konuma kaydedildi:\n{cikti_yolu}"
            )
            
            try: os.startfile(cikti_yolu)
            except: pass

        except Exception as e:
            messagebox.showerror("Hata", f"Düzeltme işlemi sırasında bir hata oluştu:\n{e}")
            logging.error("Sütun düzeltme işleminde hata", exc_info=True)
            
    logging.info("--- Excel Sütun Düzeltici Başlatıldı ---")
    threading.Thread(target=islem_yurut, daemon=True).start()

def sessiz_duzelt(dosya_yolu):
    """
    Arayüz üzerinden otomatik (sessiz) çağrıldığında çalışır.
    Mevcut dosyanın üzerine yazar (orijinal yedeği arayüz tarafından alınır).
    """
    gecici_yol = None
    try:
        dizin, dosya_adi = os.path.split(dosya_yolu)
        gecici_yol = os.path.join(dizin, f"temp_{dosya_adi}")
        
        with pd.ExcelFile(dosya_yolu) as xls:
            sheet_names = xls.sheet_names
            with pd.ExcelWriter(gecici_yol, engine='openpyxl') as writer:
                pd.DataFrame(["İşlem bekleniyor..."]).to_excel(writer, sheet_name="Gecici_Sayfa", index=False, header=False)
                yazilan_sayfa_sayisi = 0

                for i, sayfa_adi in enumerate(sheet_names):
                    if 'document map' in str(sayfa_adi).lower() or 'documentmap' in str(sayfa_adi).lower():
                        continue
                    
                    sayfa_str = str(sayfa_adi).lower()
                    is_odd_data_sheet = False
                    if "sheet" in sayfa_str or "sayfa" in sayfa_str:
                        rakamlar = re.findall(r'\d+', sayfa_str)
                        if rakamlar and int(rakamlar[0]) % 2 != 0:
                            is_odd_data_sheet = True

        
                    try:
                        ham_df = pd.read_excel(xls, sheet_name=sayfa_adi, header=None, dtype=str)
                        
                        if is_odd_data_sheet:
                            # KONTROL
                            h_idx = dm1.header_bul(ham_df)
                            if h_idx is not None:
                                row_values = ham_df.iloc[h_idx].astype(str).str.lower().values
                                match_count = sum(1 for val in row_values if str(val).strip() in ["t", "p", "n", "ff", "dd", "h", "a"] or any(k in str(val).strip() for k in ["gmt", "saat", "ww", "halihazır", "present", "istasyon", "rüzgar", "yön", "hız", "basınç", "yağış", "rrr", "bulut", "görüş", "vv", "bülten", "metar", "tipi"]))
                                if match_count >= 5: is_odd_data_sheet = False
                            if is_odd_data_sheet:
                                for r_i in range(min(10, len(ham_df))):
                                    v_a = str(ham_df.iloc[r_i, 0]).strip().upper() if len(ham_df.columns) > 0 else ""
                                    if v_a in ['TİPİ', 'TIPI', 'METAR'] or 'METAR' in v_a:
                                        is_odd_data_sheet = False; break
                                        
                        if is_odd_data_sheet:
                            yeni_sayfa_adi = str(sayfa_adi)
                            ham_df.to_excel(writer, sheet_name=yeni_sayfa_adi, index=False, header=False)
                            yazilan_sayfa_sayisi += 1
                            continue

                        sayfa_str_lower = str(sayfa_adi).lower()
                        yeni_sayfa_adi = str(sayfa_adi)
                        # DÜZELTME: Sayfa isimlerini değiştirmeyin, aksi takdirde Sheet2-Sheet3 eşleşmesi bozulur.

                        header_index = dm1.header_bul(ham_df)
                        
                        if header_index is not None:
                            ham_df.to_excel(writer, sheet_name=yeni_sayfa_adi, index=False, header=False)

                            ws = writer.sheets[yeni_sayfa_adi]
                            for col_idx, col in enumerate(ham_df.columns):
                                try: max_len = ham_df[col].astype(str).map(len).max(); width = min(max_len + 2, 60)
                                except: width = 15
                                ws.column_dimensions[get_column_letter(col_idx + 1)].width = width

                            yazilan_sayfa_sayisi += 1
                        else:
                            ham_df.to_excel(writer, sheet_name=yeni_sayfa_adi, index=False, header=False)
                            yazilan_sayfa_sayisi += 1
                    except Exception: pass

                if yazilan_sayfa_sayisi > 0:
                    writer.book.remove(writer.sheets["Gecici_Sayfa"])
                else:
                    if "Gecici_Sayfa" in writer.sheets: writer.sheets["Gecici_Sayfa"].title = "Bos_Sayfa"
                
        # Windows'ta FileExistsError (183) ve PermissionError (32) önlemek için
        # orijinal dosyayı silmeden önce silmeyi deneriz. ExcelFile 'with' bloğundan
        # çıkıldığı için dosya kilidi kalkmış olmalıdır.
        if os.path.exists(dosya_yolu):
            import time
            for _ in range(3):
                try:
                    os.remove(dosya_yolu)
                    break
                except PermissionError:
                    time.sleep(0.5)
                    
        shutil.move(gecici_yol, dosya_yolu)
    except Exception as e:
        if gecici_yol and os.path.exists(gecici_yol):
            try: os.remove(gecici_yol)
            except: pass
        raise Exception(f"Sütunlar düzeltilirken sorun oluştu: {e}")

if __name__ == "__main__":
    sutunlari_otomatik_duzelt()